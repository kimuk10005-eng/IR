import re
from pathlib import Path
from openpyxl import Workbook


TXT_DIR = Path(r"C:\Users\MASL\Desktop\txt")


def tokenize_line(line: str):
    """
    Split using:
    - tab
    - whitespace
    - ':' and '=' as extra delimiters
    """
    line = line.strip()
    if not line:
        return []
    tokens = [tok for tok in re.split(r'[\t :=]+', line) if tok]
    return tokens


def to_number(value):
    if value is None:
        return None
    m = re.fullmatch(r'[-+]?\d+(?:\.\d+)?', str(value))
    if not m:
        return None
    num = float(value)
    return int(num) if num.is_integer() else num


def extract_rawavg(tokens):
    for i, tok in enumerate(tokens):
        if tok.lower() == "rawavg" and i + 1 < len(tokens):
            return to_number(tokens[i + 1])
    return None


def extract_currentraw(tokens):
    for i, tok in enumerate(tokens):
        if tok.lower() == "currentraw" and i + 1 < len(tokens):
            return to_number(tokens[i + 1])
    return None


def build_h_series(rows, shift_rows=4):
    """
    1) Collecting baseline... RawAvg=<num>
       -> place RawAvg into H after shifting downward by 4 rows
    2) Time ... CurrentRaw <num>
       -> place CurrentRaw on the same row in H
    """
    h_values = {}

    for row_idx, tokens in rows:
        rawavg = extract_rawavg(tokens)
        if rawavg is not None:
            h_values[row_idx + shift_rows] = rawavg

    for row_idx, tokens in rows:
        current_raw = extract_currentraw(tokens)
        if current_raw is not None:
            h_values[row_idx] = current_raw

    return h_values


def auto_width(ws, max_col=8):
    for col in range(1, max_col + 1):
        max_len = 0
        col_letter = ws.cell(1, col).column_letter
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row, col).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 24)


def clear_column_h(ws):
    for r in range(1, ws.max_row + 1):
        ws.cell(r, 8).value = None


def convert_txt_to_xlsx(txt_path: Path):
    xlsx_path = txt_path.with_suffix(".xlsx")

    lines = txt_path.read_text(encoding="utf-8", errors="ignore").splitlines()
    parsed_rows = []

    for row_idx, line in enumerate(lines, start=1):
        tokens = tokenize_line(line)
        parsed_rows.append((row_idx, tokens))

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # write split tokens only to A:G
    for row_idx, tokens in parsed_rows:
        for col_idx, token in enumerate(tokens[:7], start=1):
            value = to_number(token)
            ws.cell(row=row_idx, column=col_idx, value=value if value is not None else token)

    # remove any existing values in H before inserting merged H data
    clear_column_h(ws)

    h_values = build_h_series(parsed_rows, shift_rows=4)
    for row_idx, value in h_values.items():
        ws.cell(row=row_idx, column=8, value=value)

    ws["H1"] = "Merged_H_Data"

    auto_width(ws, max_col=8)

    try:
        wb.save(xlsx_path)
        return f"[완료] {txt_path.name} -> {xlsx_path.name}"
    except PermissionError:
        return f"[스킵] 열려 있어서 저장 불가: {xlsx_path.name}"


def main():
    if not TXT_DIR.exists():
        print(f"폴더가 없습니다: {TXT_DIR}")
        return

    txt_files = sorted(TXT_DIR.glob("*.txt"))
    if not txt_files:
        print("TXT 파일이 없습니다.")
        return

    for txt_file in txt_files:
        result = convert_txt_to_xlsx(txt_file)
        print(result)


if __name__ == "__main__":
    main()
